# ============================================================
# utils/export_excel.py — Export rapport mensuel vers Excel
# ============================================================
#
# Dépendance : pip install openpyxl
# ============================================================

import os
from datetime import datetime, date

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# Couleurs Excel (ARGB)
C_BORDEAUX  = "FFC0392B"
C_OR        = "FFD4AC0D"
C_DARK      = "FF1A0A0A"
C_CARD      = "FF2A1010"
C_TEXT      = "FFF5E6D3"
C_GRIS      = "FF9E8B7A"
C_VERT      = "FF27AE60"
C_ROUGE_CLR = "FFE74C3C"
C_BLANC     = "FFFFFFFF"
C_HEADER_BG = "FF3D1515"


def _fill(hex_argb):
    return PatternFill("solid", fgColor=hex_argb)

def _font(bold=False, size=11, color=C_DARK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def _border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ============================================================
def exporter_rapport_mensuel(db, mois: str, output_path: str = None) -> str:
    """
    Génère un fichier Excel du rapport mensuel.
    mois        : "YYYY-MM"
    output_path : chemin de sortie (auto si None)
    Retourne le chemin du fichier généré.
    """
    import calendar
    from datetime import date

    if output_path is None:
        export_dir  = os.path.join(os.path.expanduser("~"), "CaveVin_Exports")
        os.makedirs(export_dir, exist_ok=True)
        output_path = os.path.join(export_dir, f"rapport_{mois}.xlsx")

    # Calcul des bornes du mois : debut = "YYYY-MM-01", fin = "YYYY-MM-31"
    annee, num_mois = int(mois.split("-")[0]), int(mois.split("-")[1])
    dernier_jour    = calendar.monthrange(annee, num_mois)[1]
    date_debut      = f"{mois}-01"
    date_fin        = f"{mois}-{dernier_jour:02d}"

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_resume_v2(wb, db, mois, date_debut, date_fin)
    _sheet_ventes_v2(wb, db, mois, date_debut, date_fin)
    _sheet_boissons_v2(wb, db, mois, date_debut, date_fin)
    _sheet_manquants_v2(wb, db, mois, date_debut, date_fin)
    _sheet_deductions(wb, db, mois)   # déductions gardent le filtre par mois texte

    wb.save(output_path)
    print(f"[EXCEL] Rapport mensuel généré : {output_path}")
    return output_path


# ---- Feuille 1 : Résumé (version corrigée avec BETWEEN) ----
def _sheet_resume_v2(wb, db, mois, date_debut, date_fin):
    ws = wb.create_sheet("Resume")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"RAPPORT MENSUEL — {mois}"
    c.font      = _font(bold=True, size=16, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — Cave OUEDRAOGO"
    c.font      = _font(size=9, color=C_GRIS, italic=True)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()

    # ── Requêtes avec BETWEEN pour éviter tout problème de DATE_FORMAT ──
    total_ventes = db.fetchone(
        "SELECT COALESCE(SUM(total),0) AS v FROM tickets "
        "WHERE date_vente BETWEEN %s AND %s AND statut='valide'",
        (date_debut, date_fin)
    ) or {"v": 0}

    nb_tickets = db.fetchone(
        "SELECT COUNT(*) AS v FROM tickets "
        "WHERE date_vente BETWEEN %s AND %s AND statut='valide'",
        (date_debut, date_fin)
    ) or {"v": 0}

    total_manquants = db.fetchone(
        "SELECT COALESCE(SUM(montant),0) AS v FROM manquants "
        "WHERE date_manquant BETWEEN %s AND %s",
        (date_debut, date_fin)
    ) or {"v": 0}

    total_deductions = db.fetchone(
        "SELECT COALESCE(SUM(montant),0) AS v FROM deductions WHERE mois=%s",
        (mois,)
    ) or {"v": 0}

    cout = db.fetchone("""
        SELECT COALESCE(SUM(tl.quantite * b.prix_achat),0) AS v
        FROM ticket_lignes tl
        JOIN boissons b ON b.id = tl.boisson_id
        JOIN tickets t  ON t.id = tl.ticket_id
        WHERE t.date_vente BETWEEN %s AND %s AND t.statut='valide'
    """, (date_debut, date_fin)) or {"v": 0}

    benefice = float(total_ventes["v"]) - float(cout["v"])

    kpis = [
        ("Chiffre d'affaires",  float(total_ventes["v"]),    "FCFA", C_OR),
        ("Nombre de tickets",   int(nb_tickets["v"]),         "",    C_TEXT),
        ("Bénéfice net",        benefice,                     "FCFA", C_VERT),
        ("Total manquants",     float(total_manquants["v"]),  "FCFA", C_ROUGE_CLR),
        ("Déductions salaires", float(total_deductions["v"]), "FCFA", C_ROUGE_CLR),
    ]

    ws.row_dimensions[3].height = 14
    row = 4
    for titre, valeur, unite, color in kpis:
        ws.merge_cells(f"A{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")
        lbl = ws[f"A{row}"]
        lbl.value     = titre
        lbl.font      = _font(bold=True, size=12, color=C_TEXT)
        lbl.fill      = _fill(C_HEADER_BG)
        lbl.alignment = _left()
        lbl.border    = _border()
        val = ws[f"D{row}"]
        disp = f"{valeur:,.0f} {unite}".strip() if isinstance(valeur, float) else f"{valeur} {unite}".strip()
        val.value     = disp
        val.font      = _font(bold=True, size=13, color=color)
        val.fill      = _fill(C_CARD)
        val.alignment = _right()
        val.border    = _border()
        ws.row_dimensions[row].height = 26
        row += 1

    for col, w in [(1,28),(2,12),(3,12),(4,18),(5,12),(6,12)]:
        _col_width(ws, col, w)


# ---- Feuille 2 : Ventes détaillées (version corrigée) ----
def _sheet_ventes_v2(wb, db, mois, date_debut, date_fin):
    ws = wb.create_sheet("Ventes")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"Ventes détaillées — {mois}"
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    hdrs = ["N° Ticket","Serveur","Date","Total (FCFA)","Montant reçu","Manquant","Statut"]
    row  = 2
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 22

    tickets = db.fetchall("""
        SELECT t.numero, CONCAT(u.prenom,' ',u.nom) AS serveur,
               t.date_vente, t.total, t.montant_recu, t.statut
        FROM tickets t
        LEFT JOIN utilisateurs u ON u.id = t.serveur_id
        WHERE t.date_vente BETWEEN %s AND %s
        ORDER BY t.date_vente, t.id
    """, (date_debut, date_fin))

    for i, r in enumerate(tickets):
        row += 1
        bg  = C_CARD if i % 2 == 0 else "FF251010"
        mq  = float(r["total"]) - float(r["montant_recu"])
        vals = [r["numero"], r["serveur"] or "—", str(r["date_vente"]),
                float(r["total"]), float(r["montant_recu"]),
                mq if mq > 0 else 0,
                r["statut"].replace("_"," ").title()]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            c.font   = _font(size=10,
                             color=C_ROUGE_CLR if col == 6 and mq > 0 else C_TEXT)
            c.alignment = _right() if col in (4,5,6) else _left()
            if col in (4,5,6) and isinstance(v, float):
                c.number_format = '#,##0'
        ws.row_dimensions[row].height = 18

    widths = [18,20,12,14,14,12,12]
    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)

    # Ligne total
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]
    c.value     = "TOTAL"
    c.font      = _font(bold=True, size=11, color=C_OR)
    c.fill      = _fill(C_HEADER_BG)
    c.alignment = _right()
    total_sum = sum(float(r["total"]) for r in tickets)
    c2 = ws.cell(row=row, column=4, value=total_sum)
    c2.font          = _font(bold=True, size=11, color=C_OR)
    c2.fill          = _fill(C_HEADER_BG)
    c2.number_format = '#,##0'
    c2.alignment     = _right()


# ---- Feuille 3 : Boissons vendues (version corrigée) ----
def _sheet_boissons_v2(wb, db, mois, date_debut, date_fin):
    ws = wb.create_sheet("Boissons")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"Boissons vendues — {mois}"
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    hdrs = ["Boisson","Catégorie","Qté vendue","CA (FCFA)","Bénéfice (FCFA)"]
    row  = 2
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 22

    rows_data = db.fetchall("""
        SELECT b.nom, b.categorie,
               SUM(tl.quantite)                              AS qte,
               SUM(tl.sous_total)                            AS ca,
               SUM(tl.quantite*(b.prix_vente-b.prix_achat)) AS benefice
        FROM ticket_lignes tl
        JOIN boissons b ON b.id = tl.boisson_id
        JOIN tickets t  ON t.id = tl.ticket_id
        WHERE t.date_vente BETWEEN %s AND %s AND t.statut='valide'
        GROUP BY b.id, b.nom, b.categorie
        ORDER BY ca DESC
    """, (date_debut, date_fin))

    for i, r in enumerate(rows_data):
        row += 1
        bg   = C_CARD if i % 2 == 0 else "FF251010"
        vals = [r["nom"], r["categorie"] or "—", int(r["qte"]),
                float(r["ca"]), float(r["benefice"] or 0)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill      = _fill(bg)
            c.border    = _border()
            c.font      = _font(size=10, color=C_TEXT)
            c.alignment = _right() if col in (3,4,5) else _left()
            if col in (3,4,5) and isinstance(v, float):
                c.number_format = '#,##0'
        ws.row_dimensions[row].height = 18

    widths = [28,14,12,14,16]
    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)

    if rows_data:
        chart = BarChart()
        chart.type        = "bar"
        chart.title       = f"CA par boisson — {mois}"
        chart.y_axis.title= "FCFA"
        chart.style       = 10
        chart.width       = 18
        chart.height      = 10
        data_ref = Reference(ws, min_col=4, min_row=2, max_row=2+len(rows_data))
        cats_ref = Reference(ws, min_col=1, min_row=3, max_row=2+len(rows_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "G3")


# ---- Feuille 4 : Manquants (version corrigée) ----
def _sheet_manquants_v2(wb, db, mois, date_debut, date_fin):
    ws = wb.create_sheet("Manquants")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"Manquants — {mois}"
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    hdrs = ["Serveur","Ticket","Montant (FCFA)","Date","Remboursé"]
    row  = 2
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 22

    rows_data = db.fetchall("""
        SELECT CONCAT(u.prenom,' ',u.nom) AS serveur, t.numero,
               m.montant, m.date_manquant, m.rembourse
        FROM manquants m
        JOIN utilisateurs u ON u.id = m.serveur_id
        LEFT JOIN tickets t ON t.id = m.ticket_id
        WHERE m.date_manquant BETWEEN %s AND %s
        ORDER BY m.date_manquant
    """, (date_debut, date_fin))

    for i, r in enumerate(rows_data):
        row += 1
        bg   = C_CARD if i % 2 == 0 else "FF251010"
        vals = [r["serveur"], r["numero"] or "—",
                float(r["montant"]), str(r["date_manquant"]),
                "Oui" if r["rembourse"] else "Non"]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill      = _fill(bg)
            c.border    = _border()
            tc = C_VERT if v == "Oui" else (C_ROUGE_CLR if v == "Non" else C_TEXT)
            c.font      = _font(size=10, color=tc)
            c.alignment = _right() if col == 3 else _left()
            if col == 3 and isinstance(v, float):
                c.number_format = '#,##0'
        ws.row_dimensions[row].height = 18

    widths = [22,18,14,14,12]
    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)


# ---- Feuille 5 : Déductions salaires ----
def _sheet_deductions(wb, db, mois):
    ws = wb.create_sheet("Deductions")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = f"Déductions salaires — {mois}"
    c.font  = _font(bold=True, size=14, color=C_OR)
    c.fill  = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    hdrs = ["Employé","Montant déduit (FCFA)","Salaire brut (FCFA)","Salaire net (FCFA)"]
    row  = 2
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 22

    rows_data = db.fetchall("""
        SELECT CONCAT(u.prenom,' ',u.nom) AS employe,
               SUM(d.montant) AS total_ded,
               u.salaire
        FROM deductions d
        JOIN utilisateurs u ON u.id=d.employe_id
        WHERE d.mois=%s
        GROUP BY d.employe_id
        ORDER BY employe
    """, (mois,))

    for i, r in enumerate(rows_data):
        row += 1
        bg   = C_CARD if i % 2 == 0 else "FF251010"
        ded  = float(r["total_ded"] or 0)
        sal  = float(r["salaire"] or 0)
        net  = sal - ded
        vals = [r["employe"], ded, sal, net]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            tc = C_VERT if col == 4 else (C_ROUGE_CLR if col == 2 else C_TEXT)
            c.font   = _font(size=10, color=tc)
            c.alignment = _right() if col > 1 else _left()
            if col > 1: c.number_format = '#,##0'
        ws.row_dimensions[row].height = 18

    widths = [24, 20, 18, 18]
    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)


# ============================================================
# Export journalier
# ============================================================
def exporter_rapport_journalier(db, jour: str, output_path: str = None) -> str:
    """
    Génère un rapport Excel pour une journée précise (YYYY-MM-DD).
    """
    if output_path is None:
        export_dir  = os.path.join(os.path.expanduser("~"), "CaveVin_Exports")
        os.makedirs(export_dir, exist_ok=True)
        output_path = os.path.join(export_dir, f"rapport_journalier_{jour}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Résumé journalier ----
    ws = wb.create_sheet("Resume")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"RAPPORT JOURNALIER — {jour}"
    c.font      = _font(bold=True, size=16, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — Cave OUEDRAOGO"
    c.font      = _font(size=9, color=C_GRIS, italic=True)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()

    total_v = db.fetchone(
        "SELECT COALESCE(SUM(total),0) AS v FROM tickets WHERE date_vente=%s AND statut='valide'",
        (jour,)) or {"v": 0}
    nb_t = db.fetchone(
        "SELECT COUNT(*) AS v FROM tickets WHERE date_vente=%s AND statut='valide'",
        (jour,)) or {"v": 0}
    manq = db.fetchone(
        "SELECT COALESCE(SUM(montant),0) AS v FROM manquants WHERE date_manquant=%s",
        (jour,)) or {"v": 0}
    cout = db.fetchone("""
        SELECT COALESCE(SUM(tl.quantite * b.prix_achat),0) AS v
        FROM ticket_lignes tl
        JOIN boissons b ON b.id=tl.boisson_id
        JOIN tickets t  ON t.id=tl.ticket_id
        WHERE t.date_vente=%s AND t.statut='valide'
    """, (jour,)) or {"v": 0}
    ben = float(total_v["v"]) - float(cout["v"])

    kpis = [
        ("Chiffre d'affaires",  float(total_v["v"]), "FCFA", C_OR),
        ("Nombre de tickets",   int(nb_t["v"]),       "",    C_TEXT),
        ("Bénéfice net",        ben,                  "FCFA", C_VERT),
        ("Total manquants",     float(manq["v"]),     "FCFA", C_ROUGE_CLR),
    ]
    row = 4
    for titre, valeur, unite, color in kpis:
        ws.merge_cells(f"A{row}:C{row}")
        ws.merge_cells(f"D{row}:F{row}")
        lbl = ws[f"A{row}"]
        lbl.value     = titre
        lbl.font      = _font(bold=True, size=12, color=C_TEXT)
        lbl.fill      = _fill(C_HEADER_BG)
        lbl.alignment = _left()
        lbl.border    = _border()
        val = ws[f"D{row}"]
        disp = f"{valeur:,.0f} {unite}".strip() if isinstance(valeur, float) else f"{valeur} {unite}".strip()
        val.value     = disp
        val.font      = _font(bold=True, size=13, color=color)
        val.fill      = _fill(C_CARD)
        val.alignment = _right()
        val.border    = _border()
        ws.row_dimensions[row].height = 26
        row += 1

    for col, w in [(1,28),(2,12),(3,12),(4,18),(5,12),(6,12)]:
        _col_width(ws, col, w)

    # ---- Tickets du jour ----
    ws2 = wb.create_sheet("Tickets du jour")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value     = f"Tickets du {jour}"
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws2.row_dimensions[1].height = 30

    hdrs   = ["N° Ticket","Serveur","Total (FCFA)","Montant reçu","Manquant","Statut","Notes"]
    widths = [18, 20, 14, 14, 12, 12, 24]
    row = 2
    for col, h in enumerate(hdrs, 1):
        c = ws2.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws2.row_dimensions[row].height = 22

    tickets = db.fetchall("""
        SELECT t.numero, CONCAT(u.prenom,' ',u.nom) AS serveur,
               t.total, t.montant_recu, t.statut, t.notes
        FROM tickets t
        LEFT JOIN utilisateurs u ON u.id=t.serveur_id
        WHERE t.date_vente=%s ORDER BY t.id
    """, (jour,))

    for i, r in enumerate(tickets):
        row += 1
        bg  = C_CARD if i % 2 == 0 else "FF251010"
        mq  = float(r["total"]) - float(r["montant_recu"])
        vals= [r["numero"], r["serveur"] or "—",
               float(r["total"]), float(r["montant_recu"]),
               mq if mq > 0 else 0,
               r["statut"].replace("_"," ").title(),
               r["notes"] or ""]
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            tc = C_ROUGE_CLR if col == 5 and mq > 0 else C_TEXT
            c.font   = _font(size=10, color=tc)
            c.alignment = _right() if col in (3,4,5) else _left()
            if col in (3,4,5) and isinstance(v, float):
                c.number_format = '#,##0'
        ws2.row_dimensions[row].height = 18

    for col, w in enumerate(widths, 1):
        _col_width(ws2, col, w)

    # ---- Boissons du jour ----
    ws3 = wb.create_sheet("Boissons")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value     = f"Boissons vendues le {jour}"
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws3.row_dimensions[1].height = 30

    hdrs   = ["Boisson","Qté","CA (FCFA)","Bénéfice (FCFA)"]
    widths = [30, 10, 14, 16]
    row = 2
    for col, h in enumerate(hdrs, 1):
        c = ws3.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()

    rows_b = db.fetchall("""
        SELECT b.nom, SUM(tl.quantite) AS qte,
               SUM(tl.sous_total) AS ca,
               SUM(tl.quantite*(b.prix_vente-b.prix_achat)) AS ben
        FROM ticket_lignes tl
        JOIN boissons b ON b.id=tl.boisson_id
        JOIN tickets t  ON t.id=tl.ticket_id
        WHERE t.date_vente=%s AND t.statut='valide'
        GROUP BY b.id ORDER BY ca DESC
    """, (jour,))

    for i, r in enumerate(rows_b):
        row += 1
        bg = C_CARD if i % 2 == 0 else "FF251010"
        vals = [r["nom"], int(r["qte"]), float(r["ca"]), float(r["ben"] or 0)]
        for col, v in enumerate(vals, 1):
            c = ws3.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            c.font   = _font(size=10, color=C_TEXT)
            c.alignment = _right() if col > 1 else _left()
            if col > 1 and isinstance(v, float): c.number_format = '#,##0'
        ws3.row_dimensions[row].height = 18

    for col, w in enumerate(widths, 1):
        _col_width(ws3, col, w)

    # ---- Manquants du jour ----
    _sheet_manquants_filtre(wb, db,
        "WHERE DATE(m.date_manquant)=%s", (jour,),
        titre=f"Manquants du {jour}")

    wb.save(output_path)
    return output_path


# ============================================================
# Export annuel
# ============================================================
def exporter_rapport_annuel(db, annee: str, output_path: str = None) -> str:
    """
    Génère un rapport Excel annuel (YYYY).
    Une ligne par mois + totaux + feuilles détail.
    """
    if output_path is None:
        export_dir  = os.path.join(os.path.expanduser("~"), "CaveVin_Exports")
        os.makedirs(export_dir, exist_ok=True)
        output_path = os.path.join(export_dir, f"rapport_annuel_{annee}.xlsx")

    wb = Workbook()
    wb.remove(wb.active)

    mois_labels = ["Janvier","Février","Mars","Avril","Mai","Juin",
                   "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    # ---- Résumé annuel par mois ----
    ws = wb.create_sheet("Resume annuel")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"RAPPORT ANNUEL — {annee}"
    c.font      = _font(bold=True, size=16, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value     = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — CaveVin Manager"
    c.font      = _font(size=9, color=C_GRIS, italic=True)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()

    # En-têtes tableau mensuel
    hdrs   = ["Mois","CA (FCFA)","Coût (FCFA)","Bénéfice (FCFA)","Tickets","Manquants (FCFA)"]
    widths = [16, 16, 14, 16, 10, 16]
    row = 4
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=11, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = 24

    total_ca = total_cout = total_ben = total_tkt = total_mq = 0

    for m in range(1, 13):
        mois_str = f"{annee}-{m:02d}"
        row += 1

        ca = db.fetchone(
            "SELECT COALESCE(SUM(total),0) AS v FROM tickets "
            "WHERE DATE_FORMAT(date_vente,'%%Y-%%m')=%s AND statut='valide'",
            (mois_str,)) or {"v": 0}
        cout = db.fetchone("""
            SELECT COALESCE(SUM(tl.quantite*b.prix_achat),0) AS v
            FROM ticket_lignes tl
            JOIN boissons b ON b.id=tl.boisson_id
            JOIN tickets t  ON t.id=tl.ticket_id
            WHERE DATE_FORMAT(t.date_vente,'%%Y-%%m')=%s AND t.statut='valide'
        """, (mois_str,)) or {"v": 0}
        nb = db.fetchone(
            "SELECT COUNT(*) AS v FROM tickets "
            "WHERE DATE_FORMAT(date_vente,'%%Y-%%m')=%s AND statut='valide'",
            (mois_str,)) or {"v": 0}
        mq = db.fetchone(
            "SELECT COALESCE(SUM(montant),0) AS v FROM manquants "
            "WHERE DATE_FORMAT(date_manquant,'%%Y-%%m')=%s",
            (mois_str,)) or {"v": 0}

        v_ca   = float(ca["v"])
        v_cout = float(cout["v"])
        v_ben  = v_ca - v_cout
        v_nb   = int(nb["v"])
        v_mq   = float(mq["v"])

        total_ca   += v_ca
        total_cout += v_cout
        total_ben  += v_ben
        total_tkt  += v_nb
        total_mq   += v_mq

        bg = C_CARD if m % 2 == 0 else "FF251010"
        vals = [mois_labels[m-1], v_ca, v_cout, v_ben, v_nb, v_mq]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            tc = C_VERT if col == 4 and v > 0 else (C_ROUGE_CLR if col == 6 and v > 0 else C_TEXT)
            c.font   = _font(size=10, color=tc)
            c.alignment = _right() if col > 1 else _left()
            if col in (2,3,4,6) and isinstance(v, float): c.number_format = '#,##0'
        ws.row_dimensions[row].height = 20

    # Ligne TOTAL
    row += 1
    totaux = ["TOTAL ANNUEL", total_ca, total_cout, total_ben, total_tkt, total_mq]
    for col, v in enumerate(totaux, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.fill   = _fill(C_HEADER_BG)
        c.border = _border()
        tc = C_OR if col == 1 else (C_VERT if col == 4 else (C_ROUGE_CLR if col == 6 else C_TEXT))
        c.font   = _font(bold=True, size=11, color=tc)
        c.alignment = _right() if col > 1 else _left()
        if col in (2,3,4,6) and isinstance(v, float): c.number_format = '#,##0'
    ws.row_dimensions[row].height = 26

    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)

    # ---- Feuilles détail (ventes + manquants année) ----
    _sheet_ventes_annee(wb, db, annee)
    _sheet_manquants_filtre(wb, db,
        "WHERE YEAR(m.date_manquant)=%s", (annee,),
        titre=f"Manquants {annee}")

    wb.save(output_path)
    return output_path


# ---- Helpers partagés ----
def _sheet_ventes_annee(wb, db, annee):
    ws = wb.create_sheet("Ventes annuelles")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"Toutes les ventes — {annee}"
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    hdrs   = ["N° Ticket","Serveur","Date","Total","Reçu","Manquant","Statut"]
    widths = [18,20,12,12,12,12,12]
    row = 2
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True, size=10, color=C_TEXT)
        c.fill = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border = _border()

    tickets = db.fetchall("""
        SELECT t.numero, CONCAT(u.prenom,' ',u.nom) AS serveur,
               t.date_vente, t.total, t.montant_recu, t.statut
        FROM tickets t
        LEFT JOIN utilisateurs u ON u.id=t.serveur_id
        WHERE YEAR(t.date_vente)=%s ORDER BY t.date_vente, t.id
    """, (annee,))

    for i, r in enumerate(tickets):
        row += 1
        bg  = C_CARD if i % 2 == 0 else "FF251010"
        mq  = float(r["total"]) - float(r["montant_recu"])
        vals= [r["numero"], r["serveur"] or "—", str(r["date_vente"]),
               float(r["total"]), float(r["montant_recu"]),
               mq if mq > 0 else 0,
               r["statut"].replace("_"," ").title()]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            c.font   = _font(size=9, color=C_ROUGE_CLR if col==6 and mq>0 else C_TEXT)
            c.alignment = _right() if col in (4,5,6) else _left()
            if col in (4,5,6) and isinstance(v, float): c.number_format = '#,##0'
        ws.row_dimensions[row].height = 16

    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)


def _sheet_manquants_filtre(wb, db, where_clause, params, titre):
    ws = wb.create_sheet("Manquants")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = titre
    c.font      = _font(bold=True, size=14, color=C_OR)
    c.fill      = _fill(C_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    hdrs   = ["Serveur","Ticket","Montant (FCFA)","Date","Remboursé"]
    widths = [22,18,14,14,12]
    row = 2
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _font(bold=True, size=10, color=C_TEXT)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = _center()
        c.border    = _border()

    rows_data = db.fetchall(f"""
        SELECT CONCAT(u.prenom,' ',u.nom) AS serveur,
               t.numero, m.montant, m.date_manquant, m.rembourse
        FROM manquants m
        JOIN utilisateurs u ON u.id=m.serveur_id
        LEFT JOIN tickets t ON t.id=m.ticket_id
        {where_clause}
        ORDER BY m.date_manquant
    """, params)

    for i, r in enumerate(rows_data):
        row += 1
        bg   = C_CARD if i % 2 == 0 else "FF251010"
        remb = "Oui" if r["rembourse"] else "Non"
        vals = [r["serveur"], r["numero"] or "—",
                float(r["montant"]), str(r["date_manquant"]), remb]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.fill   = _fill(bg)
            c.border = _border()
            tc = C_VERT if v == "Oui" else (C_ROUGE_CLR if v == "Non" else C_TEXT)
            c.font   = _font(size=10, color=tc)
            c.alignment = _right() if col == 3 else _left()
            if col == 3 and isinstance(v, float): c.number_format = '#,##0'
        ws.row_dimensions[row].height = 18

    for col, w in enumerate(widths, 1):
        _col_width(ws, col, w)
